import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

workbook = openpyxl.Workbook()
worksheet = workbook.active
for column in range(1, 6):
    for row in range(1, 6):
        worksheet.cell(row, column, row)
font = Font(name='Times New Roman', size=20, bold=True, italic=True)

for column in range(1, worksheet.max_column + 1):
    for row in range(1, worksheet.max_row + 1):
        cell = worksheet.cell(row, column)
        cell.font = font
x = 1
y = 0
for row in range(1, worksheet.max_row + 1):
    column = worksheet.max_column + x
    value = f'=SUM({get_column_letter(worksheet.min_column)}{str(row)}:{get_column_letter(worksheet.max_column + y)}{str(row)})'
    worksheet.cell(row, column, value)
    worksheet.row_dimensions[row].height = 70
    worksheet.row_dimensions[row].width = 70
    x = 0
    y = -1
worksheet.merge_cells('G1:K5')
worksheet.freeze_panes = 'A2'

ref = openpyxl.chart.Reference(worksheet, 6, 1, 6, 5)
chart = openpyxl.chart.BarChart()
chart.add_data(ref)
worksheet.add_chart(chart)
workbook.save('test.xlsx')


