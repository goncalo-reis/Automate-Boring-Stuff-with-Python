import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def mult_table(number):

    wb = openpyxl.Workbook()
    sheet = wb.active
    font = Font(bold=True)

    n = 1
    for row in range(2, number + 2):
        sheet[f'A{row}'] = n
        cell = sheet[f'A{row}']
        cell.font = font
        n +=1
    n = 1
    for column in range (2, number + 2):
        sheet[f'{get_column_letter(column)}1'] = n
        cell = sheet[f'{get_column_letter(column)}1']
        cell.font = font
        n += 1

    for column in range (2, number + 2):
        for row in range(2, number + 2):
            sheet[f'{get_column_letter(column)}{row}'] = (column - 1) * (row - 1)

    wb.save('Multiplication Table.xlsx')

mult_table(10)
