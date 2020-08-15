import openpyxl
from openpyxl.utils import get_column_letter

with open('atum.txt') as file:
    atum = file.readlines()
with open('chinelos.txt') as file:
    chinelos = file.readlines()
with open('cabelo.txt') as file:
    cabelo = file.readlines()

wb = openpyxl.Workbook()
sheet = wb.active

column = 1
row = 1
for line in atum:
    sheet[f'{get_column_letter(column)}{row}'] = line
    row += 1
column += 1
row = 1
for line in chinelos:
    sheet[f'{get_column_letter(column)}{row}'] = line
    row += 1
column += 1
row = 1
for line in cabelo:
    sheet[f'{get_column_letter(column)}{row}'] = line
    row += 1

wb.save('alright.xlsx')

