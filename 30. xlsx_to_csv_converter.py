import os, openpyxl, csv
from openpyxl.utils import get_column_letter

root = 'D:\\Ambiente de Trabalho\\excel'

for file in os.listdir(root):
    if file.endswith('.xlsx') == False:
        continue
    wb = openpyxl.load_workbook(os.path.join(root, file))
    for sheet_name in wb.sheetnames:
        sheet = wb.active
        data = []
        for row in range(1, sheet.max_row + 1):
            data_row = []
            for col in range(1, sheet.max_column + 1):
                cell = sheet[f'{get_column_letter(col)}{row}'].value
                data_row.append(cell)
            data.append(data_row)
        csv_name = f"{file.split('.')[0]}_{sheet_name}.csv"
        with open(os.path.join(root, csv_name), 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)
            csv_writer.writerows(data)    
            
