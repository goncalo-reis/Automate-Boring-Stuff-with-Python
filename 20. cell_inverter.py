import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('produceSales.xlsx', data_only=True)
sheet = wb.active
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

for column in range(1, sheet.max_column + 1):
    for row in range(1, 50):
        value = sheet[f'{get_column_letter(column)}{row}'].value
        new_sheet[f'{get_column_letter(row)}{column}'] = value

new_wb.save('hope_it_works.xlsx')

        
