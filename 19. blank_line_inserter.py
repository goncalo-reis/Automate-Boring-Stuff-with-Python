import openpyxl


def blank_line(N, M, file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    produce_list = []
    cost_per_pound_list = []
    pounds_sold_list = []
    total_list = []

    for row in range(1, N):
        produce = sheet[f'A{row}'].value
        produce_list.append(produce)
        cost_per_pound = sheet[f'B{row}'].value
        cost_per_pound_list.append(cost_per_pound)
        pounds_sold = sheet[f'C{row}'].value
        pounds_sold_list.append(pounds_sold)
        total = sheet[f'D{row}'].value
        total_list.append(total)

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active

    row = 1
    for item in produce_list:
        sheet2[f'A{row}'] = item
        row += 1
    row = 1
    for cost in cost_per_pound_list:
        sheet2[f'B{row}'] = cost
        row += 1
    row = 1
    for sold in pounds_sold_list:
        sheet2[f'C{row}'] = sold
        row += 1
    row = 1
    for numbers in total_list:
        sheet2[f'D{row}'] = numbers
        row += 1

    produce_list = []
    cost_per_pound_list = []
    pounds_sold_list = []
    total_list = []

    for row in range(N, sheet.max_row + 1):
        produce = sheet[f'A{row}'].value
        produce_list.append(produce)
        cost_per_pound = sheet[f'B{row}'].value
        cost_per_pound_list.append(cost_per_pound)
        pounds_sold = sheet[f'C{row}'].value
        pounds_sold_list.append(pounds_sold)
        total = sheet[f'D{row}'].value
        total_list.append(total)

    row = N + M
    for item in produce_list:
        sheet2[f'A{row}'] = item
        row += 1
    row = N + M
    for cost in cost_per_pound_list:
        sheet2[f'B{row}'] = cost
        row += 1
    row = N + M
    for sold in pounds_sold_list:
        sheet2[f'C{row}'] = sold
        row += 1
    row = N + M
    for numbers in total_list:
        sheet2[f'D{row}'] = numbers
        row += 1

    wb2.save('test.xlsx')

blank_line(1, 3, 'produceSales.xlsx')
    
