import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('cucu.xlsx')
sheet = wb.active

for col in range(1, sheet.max_column + 1):
    value_list = []
    for row in range(1, sheet.max_row + 1):
        value = sheet[f'{get_column_letter(col)}{row}'].value
        value_list.append(value)
    string = '\n'.join(value_list)
    with open(f'text_file_{col}.txt', 'w') as file:
        file.write(string)
